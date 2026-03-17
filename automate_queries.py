"""
MoSPI Analytics Automation - Optimized Version
=============================================
Optimizations:
1. Single file processing (as requested).
2. Parallel Processing: Uses ThreadPoolExecutor to run multiple queries at once.
3. Fast Saving: Saves in batches to reduce disk I/O.
4. Minimal Delay: Removed artificial delays.
"""

import uuid
import time
import requests
import openpyxl
import re
import concurrent.futures
from openpyxl.styles import Font, Alignment
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────────────────────────────
# 📂 CONFIGURATION
# ─────────────────────────────────────────────────────────────────
INPUT_FILENAME = "TUS.xlsx"  # The file to process
MAX_WORKERS    = 10          # Ultra-fast processing
RETRIES        = 5           # High retry count for accuracy
# ─────────────────────────────────────────────────────────────────

BASE_URL       = "http://103.48.43.11/sql-search"
API_ENDPOINT   = f"{BASE_URL}/_dash-update-component"
CONTENT_HASH   = "1c6207da962f0e15fc578d3aaf6e9b7780a9beafbf5203f0377b3d7c58cb6cf4"

HEADER_ROW     = 1
DATA_START_ROW = 2

def build_payload(query: str, session_id: str, dataset: str) -> dict:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "output": (
            f"..messages-container.children@{CONTENT_HASH}"
            f"...sessions-store.data@{CONTENT_HASH}"
            f"...user-input.value...loading-output.children.."
        ),
        "outputs": [
            {"id": "messages-container", "property": f"children@{CONTENT_HASH}"},
            {"id": "sessions-store",     "property": f"data@{CONTENT_HASH}"},
            {"id": "user-input",         "property": "value"},
            {"id": "loading-output",     "property": "children"},
        ],
        "inputs": [
            {"id": "send-btn",   "property": "n_clicks", "value": 1},
            {"id": "user-input", "property": "n_submit",  "value": None},
        ],
        "changedPropIds": ["send-btn.n_clicks"],
        "parsedChangedPropsIds": ["send-btn.n_clicks"],
        "state": [
            {"id": "user-input",        "property": "value",  "value": query},
            {"id": "current-session-id","property": "data",   "value": session_id},
            {
                "id": "sessions-store",
                "property": "data",
                "value": {
                    session_id: {
                        "id": session_id,
                        "title": "New Chat",
                        "messages": [],
                        "created_at": now,
                    }
                },
            },
            {"id": "product-name-store", "property": "data", "value": dataset},
        ],
    }

def extract_raw_response(api_json: dict) -> dict | None:
    try:
        data = api_json["response"]["sessions-store"]["data"]
        session = next(iter(data.values()))
        for msg in reversed(session.get("messages", [])):
            if msg.get("role") == "assistant":
                return msg.get("raw_response")
    except:
        pass
    return None

def format_table(columns: list, rows: list) -> str:
    if not columns or not rows: return ""
    display_rows = rows[:50]
    header = " | ".join(str(c) for c in columns)
    lines  = [header, "-" * len(header)]
    for r in display_rows:
        lines.append(" | ".join("" if v is None else str(v) for v in r))
    if len(rows) > 50: lines.append(f"... ({len(rows) - 50} more rows)")
    lines.append(f"\nTotal rows: {len(rows)}")
    return "\n".join(lines)

def process_query(query_data, dataset):
    row_num, query = query_data
    session_id = str(uuid.uuid4())
    payload = build_payload(query, session_id, dataset)
    headers = {"Content-Type": "application/json", "Referer": BASE_URL + "/"}
    
    result = {
        "row_num": row_num,
        "sql": None,
        "result_text": None,
        "remark": None,
        "success": False
    }

    # Retry loop
    for attempt in range(RETRIES):
        try:
            with requests.Session() as http:
                resp = http.post(API_ENDPOINT, json=payload, headers=headers, timeout=60)
                resp_json = resp.json()
                raw = extract_raw_response(resp_json)
                
                if raw:
                    result["sql"] = raw.get("sql", "")
                    row_count = raw.get("row_count", 0)
                    if row_count > 0:
                        result["result_text"] = format_table(raw.get("columns", []), raw.get("rows", []))
                        result["remark"] = f"{row_count} rows returned"
                    else:
                        result["remark"] = "No data available"
                    result["success"] = True
                    return result # Success!
                else:
                    try:
                        assistant_msg = resp_json['response']['sessions-store']['data'][session_id]['messages'][-1]['content']
                        result["remark"] = "No data available" if "No data available" in assistant_msg else assistant_msg[:100]
                        if result["remark"] == "No data available":
                            result["success"] = True
                            return result
                    except:
                        result["remark"] = "API returned no data"
        except Exception as e:
            result["remark"] = f"Error: {e}"
        
        if attempt < RETRIES - 1:
            time.sleep(2 * (attempt + 1)) # Exponential backoff
    
    return result

def run():
    input_path = Path(INPUT_FILENAME)
    if not input_path.exists():
        print(f"\n[ERROR] File '{INPUT_FILENAME}' not found.")
        return

    base_stem = re.sub(r'_results$', '', input_path.stem)
    output_file = input_path.parent / f"{base_stem}_results.xlsx"
    dataset = re.sub(r'[^a-zA-Z]', '', base_stem).lower()

    print(f"\n🚀 Starting OPTIMIZED automation for: {INPUT_FILENAME}")
    print(f"📊 Dataset: '{dataset}' | ⚡ Parallel workers: {MAX_WORKERS}")

    if output_file.exists():
        print(f"🔄 Resuming from {output_file.name}")
        wb = openpyxl.load_workbook(output_file)
    else:
        print(f"📄 Loading {INPUT_FILENAME}...")
        wb = openpyxl.load_workbook(input_path)
    
    ws = wb.active

    # Detect Columns
    header_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, c).value
        if val: header_map[val.strip().lower()] = c

    col_query  = header_map.get("queries", 1)
    col_sql    = header_map.get("sql query", 6)
    col_res    = header_map.get("result", 8)
    col_remark = header_map.get("remark", 10)

    queries_to_run = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        # Skip if already has SQL or remark
        if ws.cell(r, col_sql).value or ws.cell(r, col_remark).value:
            continue
        q = ws.cell(r, col_query).value
        if q and str(q).strip():
            queries_to_run.append((r, str(q).strip()))

    total = len(queries_to_run)
    if total == 0:
        print("✅ All queries already processed.")
        return

    print(f"🔎 Found {total} pending queries. Processing in parallel...")

    # Parallel execution
    completed_count = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_row = {executor.submit(process_query, q, dataset): q for q in queries_to_run}
        
        for future in concurrent.futures.as_completed(future_to_row):
            res = future.result()
            r_num = res["row_num"]
            
            # Write results back to workbook
            if res["sql"]:
                ws.cell(r_num, col_sql).value = res["sql"]
                ws.cell(r_num, col_sql).alignment = Alignment(wrap_text=True, vertical="top")
            
            if res["result_text"]:
                ws.cell(r_num, col_res).value = res["result_text"]
                ws.cell(r_num, col_res).alignment = Alignment(wrap_text=True, vertical="top")
            
            ws.cell(r_num, col_remark).value = res["remark"]
            
            completed_count += 1
            print(f"  [{completed_count}/{total}] Row {r_num} complete: {res['remark'][:40]}...")

            # Periodic save (every 10 rows) to be safe + fast
            if completed_count % 10 == 0 or completed_count == total:
                saved = False
                while not saved:
                    try:
                        wb.save(output_file)
                        saved = True
                    except PermissionError:
                        print(f"\n⚠️  ACCESS DENIED: Please CLOSE '{output_file.name}' to save progress!")
                        time.sleep(5)

    print(f"\n🏁 FINISHED! Results saved to {output_file.name}")

if __name__ == "__main__":
    run()